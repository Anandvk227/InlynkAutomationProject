import time
import unittest

import pytest
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pageObjects.ConfigurationPage import ConfigurationPage
from pageObjects.LoginPage import LoginPage
from pageObjects.AddEmployeesPage import AddEmployeesPage
from utilities.readProperties import ReadConfig
from utilities.customLogger import LogGen
from pageObjects.randomGen import randomGen
from pageObjects.companySignUpPage import companySignUpPage
from GenericLib.BaseClass import BaseClass

class addEmployees(BaseClass):
    baseURL = ReadConfig.getApplicationURL()
    DeptName = "Emp creation QA"
    DeptDescription = "Emp creation Software Testing"

    workbook = load_workbook("TestData/LoginData.xlsx")

    # Access the active worksheet
    worksheet = workbook.active

    # username = worksheet["I2"].value
    username = worksheet["A2"].value
    password = ReadConfig.getPassword()

    workbook.close()

    logger = LogGen.loggen()

    @pytest.mark.run(order=11)
    # @pytest.mark.test
    @pytest.mark.regression

    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    def test_createEmployee_superAdmin(self):
        self.logger.info("****TC_01	Create New Employee in Super Admin and Admin Account  ****")
        first_name2 = randomGen.random_first_name()
        email = randomGen.random_email()
        phone_number = randomGen.random_phone_number()
        Emp_Id = randomGen.random_Emp_Id()

        self.logger.info("******** Generating and storing data into excel sheet ***********")
        # Load the existing workbook
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        # Update the existing cells with new data
        ws['A8'] = first_name2
        ws['B8'] = "personal" + email
        ws['D8'] = phone_number
        ws['C8'] = Emp_Id
        ws['E8'] = "emp" + email

        # Save the workbook
        wb.save("TestData/LoginData.xlsx")

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.lp.clickNewsFeed()
        self.aep = AddEmployeesPage(self.driver)
        self.aep.clickEmployeesModule()
        self.aep.clickActive()
        time.sleep(3)
        self.aep.clickNewButton()
        self.aep.setFullname(first_name2)

        self.aep.setEmail("emp" + email)
        self.aep.setPersonalEmail("personal" + email)

        self.aep.setPhoneNumber(phone_number)

        self.aep.setEmpId(Emp_Id)
        self.logger.info("**** TC_02  Started Create  New Department in Super Admin Account****")
        self.aep.clickAddDeptButton()
        self.cp = ConfigurationPage(self.driver)
        self.cp.setDepartmentName("Department " + first_name2)
        self.cp.setEnterDescription(self.DeptDescription)
        self.aep.clickDoneAddDept()
        # time.sleep(3)
        xpath = self.cp.verify_DeptCreatedSuccessful_xpath
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found company name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"company name not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_createEmployee_superAdmin.png")
            self.driver.close()
            self.driver.quit()
            assert False

        self.logger.info("****TC_03  Started Create  New Division in Super Admin Account****")
        self.aep.clickAddDivisionButton()
        self.aep.setDivisionName("Division " + first_name2)
        self.cp.setEnterDescription(self.DeptDescription)
        self.aep.clickDoneAddDept()
        # time.sleep(3)
        xpath = self.cp.verify_DivisionCreatedSuccessful_xpath
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found company name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"company name not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_createEmployee_superAdmin.png")
            self.driver.close()
            self.driver.quit()
            assert False

        self.logger.info("****TC_04  Started Create  New Designation in Super Admin Account****")
        self.aep.clickAddDesignation()
        self.aep.setDivisionName("Designation " + first_name2)
        self.cp.setEnterDescription(self.DeptDescription)
        self.aep.clickDoneAddDept()
        # time.sleep(3)

        xpath = self.cp.verify_DesignationCreatedSuccessful_xpath
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_createEmployee_superAdmin.png")
            self.driver.close()
            self.driver.quit()
            assert False

        # time.sleep(2)
        self.aep.clickCountryDD()
        # time.sleep(3)
        self.sp = companySignUpPage(self.driver)
        self.sp.clickindia()
        self.sp.clickstatedd()
        self.sp.clickTelangana()
        self.sp.clickcitydd()
        self.sp.clickHyderabad()
        time.sleep(2)
        self.aep.clickAddButton()

        xpath = "//div[contains(text(), 'Employee created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found company name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"company name not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_createEmployee_superAdmin.png")
            self.driver.close()
            self.driver.quit()
            assert False


    @pytest.mark.run(order=12)
    @pytest.mark.flaky(rerun=3, rerun_delay=3)
    @pytest.mark.regression
    def test_Employee_StatusAndRole(self):
        self.logger.info("****Started Create New Employee in Super Admin and Admin Account ****")
        first_name = randomGen.random_first_name()
        email = randomGen.random_email()
        phone_number = randomGen.random_phone_number()
        Emp_Id = randomGen.random_Emp_Id()

        self.logger.info("******** Generating and storing data into excel sheet ***********")
        # Load the existing workbook
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        # Update the existing cells with new data
        ws['A11'] = first_name
        ws['B11'] = "personal" + email
        ws['D11'] = phone_number
        ws['C11'] = Emp_Id
        ws['E11'] = "emp" + email

        # Save the workbook
        wb.save("TestData/LoginData.xlsx")

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.lp.clickNewsFeed()
        self.aep = AddEmployeesPage(self.driver)
        self.aep.clickEmployeesModule()
        self.aep.clickActive()
        time.sleep(3)
        self.aep.clickNewButton()
        self.aep.setFullname(first_name)

        self.aep.setEmail("emp" + email)
        self.aep.setPersonalEmail("personal" + email)

        self.aep.setPhoneNumber(phone_number)

        self.aep.setEmpId(Emp_Id)
        # time.sleep(2)
        self.aep.ClickDD_Dept()
        self.aep.perform_keyboard_actions()
        self.aep.ClickDD_Division()
        self.aep.perform_keyboard_actions()
        self.aep.ClickDD_Designation()
        self.aep.perform_keyboard_actions()
        # time.sleep(2)
        self.aep.clickCountryDD()
        # time.sleep(3)
        self.sp = companySignUpPage(self.driver)
        self.sp.clickindia()
        self.sp.clickstatedd()
        self.sp.clickTelangana()
        self.sp.clickcitydd()
        self.sp.clickHyderabad()
        time.sleep(1)
        self.aep.clickAddButton()
        # time.sleep(3)

        xpath = "//div[contains(text(), 'Employee created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_Employee_StatusAndRole.png")
            self.driver.close()
            self.driver.quit()
            assert False

        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        first_name = ws['A11'].value
        self.logger.info("****TC_09	Employee approve*****")
        self.aep.clickActive()
        # time.sleep(2)
        self.aep.setActiveSearchField(first_name)

        first_name_xpath = "//span[contains(text(),'" + first_name + "')]"

        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, first_name_xpath))
        )
        # element.click()

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_Employee_StatusAndRole.png")
            self.driver.close()
            assert False

        element.click()
        self.aep.ClickEmployeeStatus()
        self.aep.ClickAdminStatus()
        self.logger.info("****TC_06	Check Employee is getting admin access******" )
        self.aep.ClickGrantAdmin()
        xpath = "//div[contains(text(), '" + first_name + " is an admin now')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_Employee_StatusAndRole.png")
            self.driver.close()
            self.driver.quit()
            assert False
        self.aep.ClickAdminStatus()
        self.aep.ClickEmployeesStatus()
        self.logger.info("********TC_07	Check if we Remove admin access for Employee ***********")
        self.aep.ClickRemoveStatus()
        xpath = "//div[contains(text(), '" + first_name + " is removed as admin')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_Employee_StatusAndRole.png")
            self.driver.close()
            self.driver.quit()
            assert False

        self.aep.ClickbuttonActive()
        self.logger.info("******TC_14	Verify De-active employee******")
        self.aep.ClickbuttonDeactivate()
        self.aep.setreasonText("deleting to test the functionality")
        self.aep.ClickconfDeactivate()

        xpath = "//div[contains(text(),'Employee deactivated successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_Employee_StatusAndRole.png")
            self.driver.close()
            self.driver.quit()
            assert False
        time.sleep(1)
        self.aep.setActiveSearchField(first_name)

        if "No search results found" in self.driver.page_source:
            self.logger.info("********** Acronym creation test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Acronym creation test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_Employee_StatusAndRole.png")
            assert False

    @pytest.mark.run(order=13)
    @pytest.mark.test
    @pytest.mark.regression
    def test_Employee_StatusAdminRole(self):
        self.logger.info("****TC_06	Check Employee is getting admin access ****")
        first_name = randomGen.random_first_name()
        email = randomGen.random_email()
        phone_number = randomGen.random_phone_number()
        Emp_Id = randomGen.random_Emp_Id()

        self.logger.info("******** Generating and storing data into excel sheet ***********")
        # Load the existing workbook
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        # Update the existing cells with new data
        ws['A11'] = first_name
        ws['B11'] = "personal" + email
        ws['D11'] = phone_number
        ws['C11'] = Emp_Id
        ws['E11'] = "emp" + email

        # Save the workbook
        wb.save("TestData/LoginData.xlsx")

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.lp.clickNewsFeed()
        self.aep = AddEmployeesPage(self.driver)
        self.aep.clickEmployeesModule()
        self.aep.clickActive()
        time.sleep(2)
        self.aep.clickNewButton()
        self.aep.setFullname(first_name)

        self.aep.setEmail("emp" + email)
        self.aep.setPersonalEmail("personal" + email)

        self.aep.setPhoneNumber(phone_number)

        self.aep.setEmpId(Emp_Id)
        time.sleep(2)
        self.aep.ClickDD_Dept()
        self.aep.perform_keyboard_actions()
        self.aep.ClickDD_Division()
        self.aep.perform_keyboard_actions()
        self.aep.ClickDD_Designation()
        self.aep.perform_keyboard_actions()
        # time.sleep(2)
        self.aep.clickCountryDD()
        # time.sleep(3)
        self.sp = companySignUpPage(self.driver)
        self.sp.clickindia()
        self.sp.clickstatedd()
        self.sp.clickTelangana()
        self.sp.clickcitydd()
        self.sp.clickHyderabad()
        time.sleep(1)
        self.aep.clickAddButton()
        # time.sleep(3)

        xpath = "//div[contains(text(), 'Employee created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_Employee_StatusAdminRole.png")
            self.driver.close()
            self.driver.quit()
            assert False

        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        first_name = ws['A11'].value
        self.aep.clickActive()
        # time.sleep(2)
        self.aep.setActiveSearchField(first_name)

        xpath = "//span[contains(text(),'" + first_name + "')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_Employee_StatusAdminRole.png")
            self.driver.close()
            self.driver.quit()
            assert False

        element.click()
        self.aep.ClickEmployeeStatus()
        self.aep.ClickAdminStatus()
        self.aep.ClickGrantAdmin()
        xpath = "//div[contains(text(), '" + first_name + " is an admin now')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_Employee_StatusAdminRole.png")
            self.driver.close()
            self.driver.quit()
            assert False


    if __name__ == '__main__':
        unittest.main(verbosity=2)
