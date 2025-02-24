import re
import time
import unittest

import pytest
from openpyxl.reader.excel import load_workbook
from selenium.common import NoSuchElementException, StaleElementReferenceException, TimeoutException
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from sunithaPageObjects.MyProfile import MyprofilePage
from pageObjects.LoginPage import LoginPage
from utilities.customLogger import LogGen
from utilities.readProperties import ReadConfig
from pageObjects.companySignUpPage import companySignUpPage
from pageObjects.randomGen import randomGen
from pageObjects.AddEmployeesPage import AddEmployeesPage
from pageObjects.EmployeeModulePage import EmployeeModulePage
from selenium.webdriver.support import expected_conditions as EC
from GenericLib.BaseClass import BaseClass

class TestEmployeeSignUp(BaseClass):
    baseURL = ReadConfig.getApplicationURL()
    password = ReadConfig.getPassword()

    email = randomGen.random_email()
    first_name = randomGen.random_first_name()
    phone_number = randomGen.random_phone_number()
    email1 = randomGen.random_email()
    first_name1 = randomGen.random_first_name()

    # Load the existing workbook
    wb = load_workbook("TestData/LoginData.xlsx")

    # Select the active worksheet
    ws = wb.active

    # Update the existing cells with new data
    ws['A6'] = first_name
    ws['B6'] = email
    ws['C6'] = first_name1
    ws['D6'] = email1
    # Save the changes and close the workbook
    wb.save("TestData/LoginData.xlsx")
    wb.close()
    time.sleep(5)

    @pytest.mark.run(order=14)
    # @pytest.mark.parametrize("run_number", range(1, 2))
    @pytest.mark.regression
    # @pytest.mark.skip

    def test_EmployeeSignUpValidWithoutDomain(self):
        self.logger = LogGen.loggen()
        self.logger.info("******** Starting test_Employee Sign Up with Valid ***********")
        self.logger.info("******** User is on Login page ***********")
        email = randomGen.random_email()
        first_name = randomGen.random_first_name()
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        # Read the data from the cells
        companyName = ws['C2'].value

        # Close the workbook
        # wb.close()
        # Verify the Signup functionality. with positive data.
        self.sp = companySignUpPage(self.driver)
        self.sp.clicksignuplink()
        self.sp.ClickEmployeeSignUp()
        self.logger.info("******** user is in company signup page ***********")
        self.logger.info("******** Entering valid data into the fields ***********")
        # 1. Verify the functionality and accuracy of the Company Selection dropdown.
        self.sp.setSearchCompany(companyName)
        self.sp.ClickSelectCompany()
        self.sp.setFullName(first_name)
        self.sp.setEmail(email)
        self.sp.setPhone(self.phone_number)
        self.sp.setPassword(self.password)
        self.sp.setConfirmPassword(self.password)
        self.sp.clicktermsConditions()
        time.sleep(2)
        self.logger.info("******** Clicking on signup button ***********")
        self.logger.info("******** user navigated to enter OTP page ***********")
        self.sp.clicksignupNow()
        time.sleep(2)

        # Execute JavaScript to open a new tab
        self.driver.execute_script("window.open('about:blank', '_blank');")

        # Perform actions in the new tab (if needed)
        # For example:
        self.driver.switch_to.window(self.driver.window_handles[1])
        self.logger.info("******** Opening new url in another tab for Email OTP ***********")
        time.sleep(1)
        self.driver.get("http://mailcatch.com/en/disposable-email")
        time.sleep(1)
        yopmail = self.driver.find_element(By.XPATH, "//input[@name='box']")
        yopmail.send_keys(email + Keys.ENTER)
        time.sleep(1)

        reload_button = self.driver.find_element(By.XPATH, "//img[@title='Reload']")

        # Click the Reload button every second until the subject is displayed or a maximum time is reached
        max_wait_time = 60  # Set your maximum wait time in seconds
        start_time = time.time()

        while time.time() - start_time < max_wait_time:
            reload_button.click()

            try:
                # Check if the subject is displayed
                subject = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[@class='subject']"))
                )
                subject.click()
                break  # Break out of the loop if subject is displayed
            except StaleElementReferenceException:
                print("StaleElementReferenceException occurred. Retrying...")
                continue  # Retry the loop if StaleElementReferenceException occurs
            except TimeoutException:
                time.sleep(1)

        iframeElement = self.driver.find_element(By.ID, "emailframe")
        self.driver.switch_to.frame(iframeElement)

        # Code outside the loop will be executed after the loop or when a TimeoutException occurs
        otp = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//body"))
        )
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        confirmation_code = otp.text
        getOTP = re.search(r'\b\d+\b', confirmation_code).group()
        print(getOTP)

        # This code is for QA ENV
        otp = self.driver.find_element(By.XPATH, "//body")
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        confirmation_code = otp.text
        getOTP = re.search(r'\b\d+\b', confirmation_code).group()
        print(getOTP)

        self.logger.info("******** Switching back and entering the otp ***********")
        self.driver.switch_to.default_content()

        self.driver.switch_to.window(self.driver.window_handles[0])

        self.sp.setOtp(getOTP)

        time.sleep(2)
        self.logger.info("******** Verifying the OTP ***********")
        self.sp.clickVerifyButton()
        self.sp.clickContinueToLogin()
        # email2 = ws['B6'].value
        # self.logger.info("******** Employee Sign Up successful ***********")
        # self.logger.info("******** Entering the sig up credentials for Login ***********")
        # # Read data from specific cells
        # self.lp = LoginPage(self.driver)
        # self.lp.setUserName(email2)
        # self.lp.setPassword(self.password)
        # self.lp.clickLogin()
        # self.lp.clickcreatePost()
        # self.logger.info("******** Login successful ***********")
        # act_Text = self.lp.newsFeedText()
        #
        # if act_Text == "Create News Feed":
        #     assert True
        #     self.logger.info("********* Employee SignUp Test is Passed ***********")
        #
        # else:
        #     self.driver.save_screenshot(".\\ScreenShots\\" + "test_EmployeeSignUpwithValid.png")
        #     self.logger.error("********* Employee SignUp Test is Failed ***********")
        #     self.driver.close()
        #     assert False
        #
        # time.sleep(3)
        # self.driver.find_element(By.XPATH, "//div[@class='flexAutoRow alignCntr pdngHXS']").click()
        # self.driver.close()

    @pytest.mark.run(order=15)
    @pytest.mark.regression
    def test_EmployeeSignUpWith_inValid_Data(self):
        self.logger = LogGen.loggen()
        self.logger.info("******** Starting test_Employee Sign Up with Valid ***********")
        self.logger.info("******** User is on Login page ***********")

        email = "123sdd"
        first_name = "11111111444"
        phone_number = "8888888888"
        Password = "Inlink1223"

        self.sp = companySignUpPage(self.driver)
        self.sp.clicksignuplink()
        self.sp.ClickEmployeeSignUp()
        self.logger.info("******** user is in Employee signup page ***********")
        self.logger.info("******** Entering valid data into the fields ***********")
        # Load the existing workbook
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        companyName = ws['C2'].value
        self.sp.setSearchCompany(companyName)
        self.sp.ClickSelectCompany()
        self.sp.setFullName(first_name)
        self.sp.setEmail(email)
        self.sp.setPhone(phone_number)
        self.sp.setPassword(Password)
        self.sp.setConfirmPassword(self.password)
        time.sleep(2)
        self.sp.clicktermsConditions()
        self.sp.clicktermsConditions()
        time.sleep(2)
        self.logger.info("******** Clicking on signup button ***********")
        self.sp.clicksignupNow()
        time.sleep(3)

        # List of error texts to validate
        error_texts = [
            "Please enter valid fullName",
            "Email must start with alphabet",
            "The password must be 8 characters in length, and must contain at least one small letter, one capital letter, one numeric character and one special character",
            "Password and confirm password are not same"
        ]

        # Find elements containing error texts and validate
        for text in error_texts:
            error_elements = self.driver.find_elements(By.XPATH, f"//*[contains(text(), '{text}')]")

            if error_elements:
                self.logger.info(f"Found error text: {text}")
                self.logger.info("********* Validation of Error message with invalid data test is Passed ***********")
                # self.driver.close()
                assert True

                # Additional actions can be performed here if needed
            else:
                self.logger.info(f"Error text not found: {text}")
                self.driver.save_screenshot(".\\Screenshots\\" + "test_EmployeeSignUpWith_inValid_Data.png")
                # self.logger.error("********* Validation of Error message with invalid data test is Failed ***********")
                self.driver.close()
                assert False

    @pytest.mark.run(order=18)
    @pytest.mark.regression
    # @pytest.mark.skip(reason="skip for now")
    def test_ApproveSignedUpEmployee(self):
        self.logger = LogGen.loggen()
        self.logger.info("******** Starting test_Employee Sign Up with Valid ***********")
        self.logger.info("******** User is on Login page ***********")

        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        # Read the data from the cells
        companyName = ws['C2'].value
        first_name = ws['A6'].value
        email = ws['B6'].value

        # Close the workbook
        # wb.close()
        # Verify the Signup functionality. with positive data.
        self.sp = companySignUpPage(self.driver)
        self.sp.clicksignuplink()
        self.sp.ClickEmployeeSignUp()
        self.logger.info("******** user is in company signup page ***********")
        self.logger.info("******** Entering valid data into the fields ***********")
        # 1. Verify the functionality and accuracy of the Company Selection dropdown.
        self.sp.setSearchCompany(companyName)
        self.sp.ClickSelectCompany()
        self.sp.setFullName(first_name)
        self.sp.setEmail(email)
        self.sp.setPhone(self.phone_number)
        self.sp.setPassword(self.password)
        self.sp.setConfirmPassword(self.password)
        self.sp.clicktermsConditions()
        time.sleep(2)
        self.logger.info("******** Clicking on signup button ***********")
        self.logger.info("******** user navigated to enter OTP page ***********")
        self.sp.clicksignupNow()
        time.sleep(2)

        # Execute JavaScript to open a new tab
        self.driver.execute_script("window.open('about:blank', '_blank');")

        # Perform actions in the new tab (if needed)
        # For example:
        self.driver.switch_to.window(self.driver.window_handles[1])
        self.logger.info("******** Opening new url in another tab for Email OTP ***********")
        time.sleep(1)
        self.driver.get("http://mailcatch.com/en/disposable-email")
        time.sleep(1)
        yopmail = self.driver.find_element(By.XPATH, "//input[@name='box']")
        yopmail.send_keys(email + Keys.ENTER)
        time.sleep(1)

        reload_button = self.driver.find_element(By.XPATH, "//img[@title='Reload']")

        # Click the Reload button every second until the subject is displayed or a maximum time is reached
        max_wait_time = 60  # Set your maximum wait time in seconds
        start_time = time.time()

        while time.time() - start_time < max_wait_time:
            reload_button.click()

            try:
                # Check if the subject is displayed
                subject = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[@class='subject']"))
                )
                subject.click()
                break  # Break out of the loop if subject is displayed
            except StaleElementReferenceException:
                print("StaleElementReferenceException occurred. Retrying...")
                continue  # Retry the loop if StaleElementReferenceException occurs
            except TimeoutException:
                time.sleep(1)

        iframeElement = self.driver.find_element(By.ID, "emailframe")
        self.driver.switch_to.frame(iframeElement)

        # Code outside the loop will be executed after the loop or when a TimeoutException occurs
        otp = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//body"))
        )
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        confirmation_code = otp.text
        getOTP = re.search(r'\b\d+\b', confirmation_code).group()
        print(getOTP)

        # This code is for QA ENV
        otp = self.driver.find_element(By.XPATH, "//body")
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        confirmation_code = otp.text
        getOTP = re.search(r'\b\d+\b', confirmation_code).group()
        print(getOTP)

        self.logger.info("******** Switching back and entering the otp ***********")
        self.driver.switch_to.default_content()

        self.driver.switch_to.window(self.driver.window_handles[0])

        self.sp.setOtp(getOTP)

        time.sleep(2)
        self.logger.info("******** Verifying the OTP ***********")
        self.sp.clickVerifyButton()
        self.sp.clickContinueToLogin()

        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active
        email2 = ws['A2'].value
        # first_name = ws['A6'].value

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(email2)
        self.lp.setPassword("Inlink@123")
        # self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.lp.clickNewsFeed()
        self.aep = AddEmployeesPage(self.driver)
        self.aep.clickEmployeesModule()
        self.em = EmployeeModulePage(self.driver)
        self.em.ClickPendingTab()
        self.em.setPendingSearchField(first_name)
        self.em.ClickStatusDD()
        self.em.ClickStatusApprove()
        Emp_Id = randomGen.random_Emp_Id()
        self.em.setEmpId(Emp_Id)
        self.em.ClickDepartmentDD()
        self.em.ClickSelectDD()
        self.em.ClickDivisionDD()
        self.em.ClickSelectDD()
        self.em.ClickDesignationDD()
        self.em.ClickSelectDD()
        self.em.ClickApproveButton()
        xpath = "//div[contains(text(), 'Employee approved successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Toast Message : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Toast Message not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_ApproveSignedUpEmployee.png")
            self.driver.close()
            self.driver.quit()
            assert False


        # Verify active tab
        self.logger.info("Verifying the Approved Employee in Active Tab")
        self.aep.clickActive()
        time.sleep(1)
        self.aep.setActiveSearchField(first_name)
        element = self.driver.find_element(By.XPATH, "//span[contains(text(),'" + first_name + "')]")
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()

        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_ApproveSignedUpEmployee.png")
            self.driver.close()
            assert False

        element.click()

    @pytest.mark.run(order=16)
    @pytest.mark.regression
    # @pytest.mark.tests

    # @pytest.mark.flaky(rerun=3, rerun_delay=2)
    # def test_EmployeeSignUpWithValid(self, run_number, setup):
    def test_EmployeeSignUpValidWithoutDomainAdmin(self):
        self.logger = LogGen.loggen()
        self.logger.info("******** Starting test_Employee Sign Up with Valid ***********")
        self.logger.info("******** Starting test_Employee Sign Up with Valid ***********")
        self.logger.info("******** User is on Login page ***********")


        phone_number = randomGen.random_phone_number()

        self.logger.info("******** Generating and storing data into excel sheet ***********")
        # Load the existing workbook
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        companyName = ws['C2'].value
        first_name1 = ws['C6'].value
        email1 = ws['D6'].value
        companyEmail = ws['A2'].value

        # Update the existing cells with new data

        # ws['C6'] = phone_number

        # Save the workbook
        wb.save("TestData/LoginData.xlsx")
        # Verify the Signup functionality. with positive data.
        self.sp = companySignUpPage(self.driver)
        self.sp.clicksignuplink()
        self.sp.ClickEmployeeSignUp()
        self.logger.info("******** user is in company signup page ***********")
        self.logger.info("******** Entering valid data into the fields ***********")
        # 1. Verify the functionality and accuracy of the Company Selection dropdown.
        self.sp.setSearchCompany(companyName)
        self.sp.ClickSelectCompany()
        self.sp.setFullName(first_name1)
        self.sp.setEmail(email1)
        self.sp.setPhone(phone_number)
        self.sp.setPassword(self.password)
        self.sp.setConfirmPassword(self.password)
        self.sp.clicktermsConditions()
        time.sleep(2)
        self.logger.info("******** Clicking on signup button ***********")
        self.logger.info("******** user navigated to enter OTP page ***********")
        self.sp.clicksignupNow()
        time.sleep(2)

        # Execute JavaScript to open a new tab
        self.driver.execute_script("window.open('about:blank', '_blank');")

        # Perform actions in the new tab (if needed)
        # For example:
        self.driver.switch_to.window(self.driver.window_handles[1])
        self.logger.info("******** Opening new url in another tab for Email OTP ***********")
        time.sleep(1)
        self.driver.get("http://mailcatch.com/en/disposable-email")
        time.sleep(1)
        yopmail = self.driver.find_element(By.XPATH, "//input[@name='box']")
        yopmail.send_keys(email1 + Keys.ENTER)
        time.sleep(1)

        reload_button = self.driver.find_element(By.XPATH, "//img[@title='Reload']")

        # Click the Reload button every second until the subject is displayed or a maximum time is reached
        max_wait_time = 60  # Set your maximum wait time in seconds
        start_time = time.time()

        while time.time() - start_time < max_wait_time:
            reload_button.click()

            try:
                # Check if the subject is displayed
                subject = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[@class='subject']"))
                )
                subject.click()
                break  # Break out of the loop if subject is displayed
            except StaleElementReferenceException:
                print("StaleElementReferenceException occurred. Retrying...")
                continue  # Retry the loop if StaleElementReferenceException occurs
            except TimeoutException:
                time.sleep(1)

        iframeElement = self.driver.find_element(By.ID, "emailframe")
        self.driver.switch_to.frame(iframeElement)

        # Code outside the loop will be executed after the loop or when a TimeoutException occurs
        otp = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//body"))
        )
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        confirmation_code = otp.text
        getOTP = re.search(r'\b\d+\b', confirmation_code).group()
        print(getOTP)

        # This code is for QA ENV
        otp = self.driver.find_element(By.XPATH, "//body")
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        confirmation_code = otp.text
        getOTP = re.search(r'\b\d+\b', confirmation_code).group()
        print(getOTP)

        self.logger.info("******** Switching back and entering the otp ***********")
        self.driver.switch_to.default_content()

        self.driver.switch_to.window(self.driver.window_handles[0])

        self.sp.setOtp(getOTP)

        time.sleep(2)
        self.logger.info("******** Verifying the OTP ***********")
        self.sp.clickVerifyButton()
        self.sp.clickContinueToLogin()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(companyEmail)
        self.lp.setPassword("Inlink@123")
        # self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.lp.clickNewsFeed()
        self.aep = AddEmployeesPage(self.driver)
        self.aep.clickEmployeesModule()
        self.em = EmployeeModulePage(self.driver)
        self.em.ClickPendingTab()
        self.em.setPendingSearchField(first_name1)
        self.em.ClickStatusDD()
        self.em.ClickStatusApprove()
        Emp_Id = randomGen.random_Emp_Id()
        self.em.setEmpId(Emp_Id)
        self.em.ClickDepartmentDD()
        self.em.ClickSelectDD()
        self.em.ClickDivisionDD()
        self.em.ClickSelectDD()
        self.em.ClickDesignationDD()
        self.em.ClickSelectDD()
        self.em.ClickApproveButton()
        xpath = "//div[contains(text(), 'Employee approved successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Toast Message : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Toast Message not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_EmployeeSignUpValidWithoutDomainAdmin.png")
            self.driver.close()
            self.driver.quit()
            assert False

        # Verify active tab
        self.logger.info("Verifying the Approved Employee in Active Tab")
        self.aep.clickActive()
        time.sleep(1)
        self.aep.setActiveSearchField(first_name1)
        element_xpath = "//span[contains(text(),'" + first_name1 + "')]"
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, element_xpath))
        )
        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()

        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_EmployeeSignUpValidWithoutDomainAdmin.png")
            # self.driver.close()
            assert False
        element.click()
        self.aep.ClickEmployeeStatus()
        self.aep.ClickAdminStatus()
        self.aep.ClickGrantAdmin()

        xpath = "//div[contains(text(), '" + first_name1 + " is an admin now')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_EmployeeSignUpValidWithoutDomainAdmin.png")
            # self.driver.close()
            # self.driver.quit()
            assert False

    @pytest.mark.run(order=17)
    @pytest.mark.regression
    @pytest.mark.testing
    # @pytest.mark.tests
    def test_RejectSignedUpEmployee(self):
        self.logger = LogGen.loggen()
        self.logger.info("******** Starting test_Employee Sign Up with Valid ***********")
        self.logger.info("******** User is on Login page ***********")
        email = randomGen.random_email()
        first_name = randomGen.random_first_name()
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        # Read the data from the cells
        companyName = ws['C2'].value

        # Close the workbook
        # wb.close()
        # Verify the Signup functionality. with positive data.
        self.sp = companySignUpPage(self.driver)
        self.sp.clicksignuplink()
        self.sp.ClickEmployeeSignUp()
        self.logger.info("******** user is in company signup page ***********")
        self.logger.info("******** Entering valid data into the fields ***********")
        # 1. Verify the functionality and accuracy of the Company Selection dropdown.
        self.sp.setSearchCompany(companyName)
        self.sp.ClickSelectCompany()
        self.sp.setFullName(first_name)
        self.sp.setEmail(email)
        self.sp.setPhone(self.phone_number)
        self.sp.setPassword(self.password)
        self.sp.setConfirmPassword(self.password)
        self.sp.clicktermsConditions()
        time.sleep(2)
        self.logger.info("******** Clicking on signup button ***********")
        self.logger.info("******** user navigated to enter OTP page ***********")
        self.sp.clicksignupNow()
        time.sleep(2)

        # Execute JavaScript to open a new tab
        self.driver.execute_script("window.open('about:blank', '_blank');")

        # Perform actions in the new tab (if needed)
        # For example:
        self.driver.switch_to.window(self.driver.window_handles[1])
        self.logger.info("******** Opening new url in another tab for Email OTP ***********")
        time.sleep(1)
        self.driver.get("http://mailcatch.com/en/disposable-email")
        time.sleep(1)
        yopmail = self.driver.find_element(By.XPATH, "//input[@name='box']")
        yopmail.send_keys(email + Keys.ENTER)
        time.sleep(1)

        reload_button = self.driver.find_element(By.XPATH, "//img[@title='Reload']")

        # Click the Reload button every second until the subject is displayed or a maximum time is reached
        max_wait_time = 60  # Set your maximum wait time in seconds
        start_time = time.time()

        while time.time() - start_time < max_wait_time:
            reload_button.click()

            try:
                # Check if the subject is displayed
                subject = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[@class='subject']"))
                )
                subject.click()
                break  # Break out of the loop if subject is displayed
            except StaleElementReferenceException:
                print("StaleElementReferenceException occurred. Retrying...")
                continue  # Retry the loop if StaleElementReferenceException occurs
            except TimeoutException:
                time.sleep(1)

        iframeElement = self.driver.find_element(By.ID, "emailframe")
        self.driver.switch_to.frame(iframeElement)

        # Code outside the loop will be executed after the loop or when a TimeoutException occurs
        otp = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//body"))
        )
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        confirmation_code = otp.text
        getOTP = re.search(r'\b\d+\b', confirmation_code).group()
        print(getOTP)

        # This code is for QA ENV
        otp = self.driver.find_element(By.XPATH, "//body")
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        confirmation_code = otp.text
        getOTP = re.search(r'\b\d+\b', confirmation_code).group()
        print(getOTP)

        self.logger.info("******** Switching back and entering the otp ***********")
        self.driver.switch_to.default_content()

        self.driver.switch_to.window(self.driver.window_handles[0])

        self.sp.setOtp(getOTP)

        time.sleep(2)
        self.logger.info("******** Verifying the OTP ***********")
        self.sp.clickVerifyButton()
        self.sp.clickContinueToLogin()
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active
        email = ws['A2'].value
        # first_name = ws['A6'].value
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(email)
        self.lp.setPassword("Inlink@123")
        # self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.lp.clickNewsFeed()
        self.aep = AddEmployeesPage(self.driver)
        self.aep.clickEmployeesModule()
        self.em = EmployeeModulePage(self.driver)
        self.em.ClickPendingTab()
        self.em.setPendingSearchField(first_name)
        self.em.ClickStatusDD()
        self.em.ClickStatusReject()
        self.em.ClickConfReject()
        time.sleep(5)
        Verify_texts = [
            "Employee rejected successfully"
            # Add more error texts if needed
        ]

        # Find elements containing error texts and validate
        for text in Verify_texts:
            error_elements = self.driver.find_elements(By.XPATH, f"//*[contains(text(), '{text}')]")

            if error_elements:
                self.logger.info(f"Found Toast Message: {text}")
                assert True

            else:
                self.logger.info(f"Toast Message not found: {text}")
                self.driver.save_screenshot(".\\Screenshots\\" + "test_RejectSignedUpEmployee.png")
                # self.logger.error("********* Validation of Error message with invalid data test is Failed ***********")
                # self.driver.close()
                assert False

        # Verify active tab
        self.logger.info("Verifying the rejected Employee in rejected Tab")
        self.em.ClickRejectTab()
        time.sleep(1)
        self.em.setRejectedSearchField(first_name)
        element = self.driver.find_element(By.XPATH, "//span[contains(text(),'" + first_name + "')]")


        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_RejectSignedUpEmployee.png")
            self.driver.close()
            assert False

    # @pytest.mark.test
    # change credentials to run
    # @pytest.mark.flaky(rerun=35, rerun_delay=3)
    @pytest.mark.regression
    @pytest.mark.run(order=19)
    def test_EmployeeSignUpValidWithDomain(self):
        # for _ in range(30):
            self.logger = LogGen.loggen()
            self.logger.info("******** Starting test_Employee Sign Up Valid With domain***********")
            self.logger.info("******** User is on Login page ***********")

            # email = randomGen.random_email()
            first_name = randomGen.random_first_name()
            first_name1 = randomGen.random_Ranfirst_name()
            phone_number = randomGen.random_phone_number()

            self.logger.info("******** Generating and storing data into excel sheet ***********")
            # Load the existing workbook
            wb = load_workbook("TestData/LoginData.xlsx")

            # Select the active worksheet
            ws = wb.active

            companyName = ws['A10'].value
            companyEmail = ws['B10'].value

            # Update the existing cells with new data
            # ws['E10'] = email
            ws['C10'] = first_name
            ws['D10'] = phone_number

            # Save the workbook
            wb.save("TestData/LoginData.xlsx")
            # Verify the Signup functionality. with positive data.
            self.sp = companySignUpPage(self.driver)
            self.sp.clicksignuplink()
            self.sp.ClickEmployeeSignUp()
            self.logger.info("******** user is in company signup page ***********")
            self.logger.info("******** Entering valid data into the fields ***********")
            # 1. Verify the functionality and accuracy of the Company Selection dropdown.
            self.sp.setSearchCompany(companyName)
            self.sp.ClickSelectCompany()
            self.sp.setFullName(first_name)
            time.sleep(0.5)

            self.sp.setEmail(first_name1 + "20")
            Email = first_name1 + "20@mailcatch.com"
            ws['E10'] = Email
            wb.save("TestData/LoginData.xlsx")
            time.sleep(1)
            self.sp.setPhone(phone_number)
            self.sp.setPassword(self.password)
            self.sp.setConfirmPassword(self.password)
            self.sp.clicktermsConditions()
            time.sleep(2)
            self.logger.info("******** Clicking on signup button ***********")
            self.logger.info("******** user navigated to enter OTP page ***********")
            self.sp.clicksignupNow()
            time.sleep(2)


            # Execute JavaScript to open a new tab
            self.driver.execute_script("window.open('about:blank', '_blank');")

            # Perform actions in the new tab (if needed)
            # For example:
            self.driver.switch_to.window(self.driver.window_handles[1])
            self.logger.info("******** Opening new url in another tab for Email OTP ***********")
            time.sleep(1)
            self.driver.get("http://mailcatch.com/en/disposable-email")
            time.sleep(1)
            yopmail = self.driver.find_element(By.XPATH, "//input[@name='box']")
            yopmail.send_keys(Email + Keys.ENTER)
            time.sleep(1)

            reload_button = self.driver.find_element(By.XPATH, "//img[@title='Reload']")

            # Click the Reload button every second until the subject is displayed or a maximum time is reached
            max_wait_time = 60  # Set your maximum wait time in seconds
            start_time = time.time()

            while time.time() - start_time < max_wait_time:
                reload_button.click()

                try:
                    # Check if the subject is displayed
                    subject = WebDriverWait(self.driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, "//td[@class='subject']"))
                    )
                    subject.click()
                    break  # Break out of the loop if subject is displayed
                except StaleElementReferenceException:
                    print("StaleElementReferenceException occurred. Retrying...")
                    continue  # Retry the loop if StaleElementReferenceException occurs
                except TimeoutException:
                    time.sleep(1)

            iframeElement = self.driver.find_element(By.ID, "emailframe")
            self.driver.switch_to.frame(iframeElement)

            # Code outside the loop will be executed after the loop or when a TimeoutException occurs
            otp = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//body"))
            )
            self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
            time.sleep(0.5)

            confirmation_code = otp.text
            getOTP = re.search(r'\b\d+\b', confirmation_code).group()
            print(getOTP)

            # This code is for QA ENV
            otp = self.driver.find_element(By.XPATH, "//body")
            self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
            time.sleep(0.5)

            confirmation_code = otp.text
            getOTP = re.search(r'\b\d+\b', confirmation_code).group()
            print(getOTP)

            self.logger.info("******** Switching back and entering the otp ***********")
            self.driver.switch_to.default_content()

            self.driver.switch_to.window(self.driver.window_handles[0])

            self.sp.setOtp(getOTP)

            time.sleep(2)
            self.logger.info("******** Verifying the OTP ***********")
            self.sp.clickVerifyButton()
            self.sp.clickContinueToLogin()

        # self.driver.close()

    @pytest.mark.run(order=20)
    @pytest.mark.regression
    # @pytest.mark.test
    # @pytest.mark.flaky(rerun=3, rerun_delay=2)
    # Change credentials to run
    def test_ActiveSignedUpEmployeeWithDomain(self):
        # to use same class different test method
        self.test_EmployeeSignUpValidWithDomain()

        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active
        email = ws['B10'].value
        first_name = ws['C10'].value
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(email)
        self.lp.setPassword("Inlink@123")
        # self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.lp.clickNewsFeed()
        self.aep = AddEmployeesPage(self.driver)
        self.aep.clickEmployeesModule()
        self.em = EmployeeModulePage(self.driver)

        self.aep.clickActive()

        self.aep.setActiveSearchField(first_name)

        element_xpath = "//span[contains(text(),'" + first_name + "')]"
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, element_xpath))
        )

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_ActiveSignedUpEmployeeWithDomain.png")
            self.driver.close()
            assert False

    @pytest.mark.run(order=21)
    @pytest.mark.regression
    # @pytest.mark.test
    def test_AdminSignedUpEmployeeWithDomain(self):
        # to use same class different test method
        self.test_EmployeeSignUpValidWithDomain()

        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active
        email = ws['B10'].value
        first_name = ws['C10'].value
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(email)
        self.lp.setPassword("Inlink@123")
        # self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.lp.clickNewsFeed()
        self.aep = AddEmployeesPage(self.driver)
        self.aep.clickEmployeesModule()
        self.em = EmployeeModulePage(self.driver)

        self.aep.clickActive()

        self.aep.setActiveSearchField(first_name)

        element_xpath = "//span[contains(text(),'" + first_name + "')]"
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, element_xpath))
        )

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_AdminSignedUpEmployeeWithDomain.png")
            self.driver.close()
            assert False

        element.click()
        self.aep.ClickEmployeeStatus()
        self.aep.ClickAdminStatus()
        self.aep.ClickGrantAdmin()
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active
        first_name = ws['C10'].value

        xpath = "//div[contains(text(), '" + first_name + " is an admin now')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_AdminSignedUpEmployeeWithDomain.png")
            self.driver.close()
            self.driver.quit()
            assert False

        self.lp.clickLogout()
        self.logger.info("******** Entering Admin Employee credentials for verification ***********")
        # Read data from specific cells
        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        Username = ws['E10'].value
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(Username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.my = MyprofilePage(self.driver)
        self.my.clickMyProfileModule()

        adminxpath = "// span[text() = 'admin']"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, adminxpath))
        )

        if element:
            self.logger.info(f"Found Employee Role : {element.text}")
            assert True

        else:
            self.logger.info(f"Employee Role not found: {element.text}")
            self.driver.save_screenshot(".\\Screenshots\\" + "VerifyAdminRole.png")
            self.driver.close()
            self.driver.quit()
            assert False

if __name__ == "__main__":
    unittest.main()

