import re
import time

import pytest
from openpyxl.reader.excel import load_workbook
from selenium.common import StaleElementReferenceException, TimeoutException
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from pageObjects.companySignUpPage import companySignUpPage
from pageObjects.AddEmployeesPage import AddEmployeesPage
from pageObjects.EmployeeModulePage import EmployeeModulePage
from utilities.readProperties import ReadConfig
from GenericLib.BaseClass import BaseClass
from utilities.customLogger import LogGen
from pageObjects.LoginPage import LoginPage
from pageObjects.randomGen import randomGen
from selenium.webdriver.support import expected_conditions as EC


class addEmployees(BaseClass):
    baseURL = ReadConfig.getApplicationURL()
    workbook = load_workbook("TestData/LoginData.xlsx")

    # Access the active worksheet
    worksheet = workbook.active

    # username = worksheet["I2"].value
    username = worksheet["A2"].value
    password = ReadConfig.getPassword()

    workbook.close()

    logger = LogGen.loggen()

    @pytest.mark.run(order=1)
    # @pytest.mark.skip
    @pytest.mark.regression
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    def test_InviteEmployeeCreation(self):
        self.logger.info("****TC_01	Create New Employee in Super Admin and Admin Account  ****")
        first_name = randomGen.random_first_name()
        email = randomGen.random_email()
        email2 = randomGen.random_email()
        phone_number = randomGen.random_phone_number()
        Emp_Id = randomGen.random_Emp_Id()

        self.logger.info("******** Generating and storing data into excel sheet ***********")
        # Load the existing workbook
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        # Update the existing cells with new data
        ws['A18'] = first_name
        ws['B18'] = "personal" + email
        ws['D18'] = phone_number
        ws['C18'] = email
        ws['E18'] = email2

        # Save the workbook
        wb.save("TestData/LoginData.xlsx")

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.lp.clickNewsFeed()
        self.aep = AddEmployeesPage(self.driver)
        self.aep.clickEmployeesModule()
        self.aep.ClickcInviteButton()
        self.aep.ClickDDcompany()
        self.aep.selctEmployee()
        self.aep.ClickcCopyButton()
        xpath = "//div[contains(text(),'Copied successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_InviteEmployeeCreation.png")
            assert False

        self.aep.setInviteEmail(email)
        self.aep.ClickcSubmitButton()
        self.aep.ClickcSendButton()

        xpath = "//div[contains(text(),'Invitation send successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_InviteEmployeeCreation.png")
            assert False

        self.lp.clickLogout()

        # Execute JavaScript to open a new tab
        self.driver.execute_script("window.open('about:blank', '_blank');")

        # Perform actions in the new tab (if needed)
        # For example:
        self.driver.switch_to.window(self.driver.window_handles[1])
        self.logger.info("******** Opening new url in another tab for Email OTP ***********")
        time.sleep(1)
        self.driver.get("http://mailcatch.com/en/disposable-email")
        time.sleep(1)
        yopmail = self.driver.find_element(By.XPATH, "//input[@name='box']")
        yopmail.send_keys(email + Keys.ENTER)
        time.sleep(1)

        reload_button = self.driver.find_element(By.XPATH, "//img[@title='Reload']")

        # Click the Reload button every second until the subject is displayed or a maximum time is reached
        max_wait_time = 60  # Set your maximum wait time in seconds
        start_time = time.time()

        while time.time() - start_time < max_wait_time:
            reload_button.click()

            try:
                # Check if the subject is displayed
                subject = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[@class='subject']"))
                )
                subject.click()
                break  # Break out of the loop if subject is displayed
            except StaleElementReferenceException:
                print("StaleElementReferenceException occurred. Retrying...")
                continue  # Retry the loop if StaleElementReferenceException occurs
            except TimeoutException:
                time.sleep(1)

        iframeElement = self.driver.find_element(By.ID, "emailframe")
        self.driver.switch_to.frame(iframeElement)

        # Code outside the loop will be executed after the loop or when a TimeoutException occurs
        otp = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//body"))
        )
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        # This code is for QA ENV
        otp = self.driver.find_element(By.XPATH, "//body")
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        Url = otp.text

        self.logger.info("******** Switching back and entering the otp ***********")
        self.driver.switch_to.default_content()

        self.driver.switch_to.window(self.driver.window_handles[0])
        self.driver.get(Url)

        self.sp = companySignUpPage(self.driver)
        time.sleep(5)
        self.sp.setFullName(first_name)
        self.sp.setEmail(email2)
        self.sp.setPhone(phone_number)
        self.sp.setPassword(self.password)
        self.sp.setConfirmPassword(self.password)
        self.sp.clicktermsConditions()
        time.sleep(2)
        self.logger.info("******** Clicking on signup button ***********")
        self.logger.info("******** user navigated to enter OTP page ***********")
        self.sp.clicksignupNow()
        time.sleep(2)

        # Execute JavaScript to open a new tab
        self.driver.execute_script("window.open('about:blank', '_blank');")

        # Perform actions in the new tab (if needed)
        # For example:
        self.driver.switch_to.window(self.driver.window_handles[1])
        self.logger.info("******** Opening new url in another tab for Email OTP ***********")
        time.sleep(1)
        self.driver.get("http://mailcatch.com/en/disposable-email")
        time.sleep(1)
        yopmail = self.driver.find_element(By.XPATH, "//input[@name='box']")
        yopmail.send_keys(email2 + Keys.ENTER)
        time.sleep(1)

        reload_button = self.driver.find_element(By.XPATH, "//img[@title='Reload']")

        # Click the Reload button every second until the subject is displayed or a maximum time is reached
        max_wait_time = 60  # Set your maximum wait time in seconds
        start_time = time.time()

        while time.time() - start_time < max_wait_time:
            reload_button.click()

            try:
                # Check if the subject is displayed
                subject = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[@class='subject']"))
                )
                subject.click()
                break  # Break out of the loop if subject is displayed
            except StaleElementReferenceException:
                print("StaleElementReferenceException occurred. Retrying...")
                continue  # Retry the loop if StaleElementReferenceException occurs
            except TimeoutException:
                time.sleep(1)

        iframeElement = self.driver.find_element(By.ID, "emailframe")
        self.driver.switch_to.frame(iframeElement)

        # Code outside the loop will be executed after the loop or when a TimeoutException occurs
        otp = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//body"))
        )
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        confirmation_code = otp.text
        getOTP = re.search(r'\b\d+\b', confirmation_code).group()
        print(getOTP)

        # This code is for QA ENV
        otp = self.driver.find_element(By.XPATH, "//body")
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        confirmation_code = otp.text
        getOTP = re.search(r'\b\d+\b', confirmation_code).group()
        print(getOTP)

        self.logger.info("******** Switching back and entering the otp ***********")
        self.driver.switch_to.default_content()

        self.driver.switch_to.window(self.driver.window_handles[0])

        self.sp.setOtp(getOTP)

        time.sleep(2)
        self.logger.info("******** Verifying the OTP ***********")
        self.sp.clickVerifyButton()
        self.sp.clickContinueToLogin()

        self.lp.setUserName(self.username)
        self.lp.setPassword("Inlink@123")
        # self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.lp.clickNewsFeed()
        self.aep = AddEmployeesPage(self.driver)
        self.aep.clickEmployeesModule()
        self.em = EmployeeModulePage(self.driver)
        self.em.ClickPendingTab()
        self.em.setPendingSearchField(first_name)
        self.em.ClickStatusDD()
        self.em.ClickStatusApprove()
        Emp_Id = randomGen.random_Emp_Id()
        self.em.setEmpId(Emp_Id)
        self.em.ClickDepartmentDD()
        self.em.ClickSelectDD()
        self.em.ClickDivisionDD()
        self.em.ClickSelectDD()
        self.em.ClickDesignationDD()
        self.em.ClickSelectDD()
        self.em.ClickApproveButton()
        xpath = "//div[contains(text(), 'Employee approved successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Toast Message : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Toast Message not found: {element.text}")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_ApproveSignedUpEmployee.png")
            self.driver.close()
            self.driver.quit()
            assert False

        self.lp.clickLogout()
        self.logger.info("******** Employee Sign Up successful ***********")
        self.logger.info("******** Entering the sig up credentials for Login ***********")
        # Read data from specific cells
        email = ws['E18'].value

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(email)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.lp.clickcreatePost()

    @pytest.mark.run(order=2)
    # @pytest.mark.test
    @pytest.mark.regression
    # @pytest.mark.flaky(reruns=3, reruns_delay=2)
    def test_InviteCompanyCreation(self):
        self.logger.info("****TC_01	Create New Employee in Super Admin and Admin Account  ****")
        baseURL = ReadConfig.getApplicationURL()
        setSearchIndustryType = "Information Technology"
        password = ReadConfig.getPassword()

        email = randomGen.random_email()
        email2 = randomGen.random_email()
        first_name = randomGen.random_first_name()
        company_name = randomGen.random_company_name()
        phone_number = randomGen.random_phone_number()
        # Load the existing workbook
        workbook = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        worksheet = workbook.active

        # Update the existing cells with new data
        worksheet['B19'] = email
        worksheet['E19'] = email2
        worksheet['B19'] = first_name
        worksheet['A19'] = company_name
        worksheet['D19'] = phone_number
        # Save the workbook
        workbook.save("TestData/LoginData.xlsx")

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.lp.clickNewsFeed()
        self.aep = AddEmployeesPage(self.driver)
        self.aep.clickEmployeesModule()
        self.aep.ClickcInviteButton()
        self.aep.ClickDDcompany()
        self.aep.selctCompany()
        self.aep.ClickcCopyButton()
        xpath = "//div[contains(text(),'Copied successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_InviteCompanyCreation.png")
            assert False

        self.aep.setInviteEmail(email)
        self.aep.ClickcSubmitButton()
        self.aep.ClickcSendButton()

        xpath = "//div[contains(text(),'Invitation send successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_InviteCompanyCreation.png")
            assert False

        self.lp.clickLogout()

        # Execute JavaScript to open a new tab
        self.driver.execute_script("window.open('about:blank', '_blank');")

        # Perform actions in the new tab (if needed)
        # For example:
        self.driver.switch_to.window(self.driver.window_handles[1])
        self.logger.info("******** Opening new url in another tab for Email OTP ***********")
        time.sleep(1)
        self.driver.get("http://mailcatch.com/en/disposable-email")
        time.sleep(1)
        yopmail = self.driver.find_element(By.XPATH, "//input[@name='box']")
        yopmail.send_keys(email + Keys.ENTER)
        time.sleep(1)

        reload_button = self.driver.find_element(By.XPATH, "//img[@title='Reload']")

        # Click the Reload button every second until the subject is displayed or a maximum time is reached
        max_wait_time = 60  # Set your maximum wait time in seconds
        start_time = time.time()

        while time.time() - start_time < max_wait_time:
            reload_button.click()

            try:
                # Check if the subject is displayed
                subject = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[@class='subject']"))
                )
                subject.click()
                break  # Break out of the loop if subject is displayed
            except StaleElementReferenceException:
                print("StaleElementReferenceException occurred. Retrying...")
                continue  # Retry the loop if StaleElementReferenceException occurs
            except TimeoutException:
                time.sleep(1)

        iframeElement = self.driver.find_element(By.ID, "emailframe")
        self.driver.switch_to.frame(iframeElement)

        # Code outside the loop will be executed after the loop or when a TimeoutException occurs
        otp = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//body"))
        )
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        # This code is for QA ENV
        otp = self.driver.find_element(By.XPATH, "//body")
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        Url = otp.text

        self.logger.info("******** Switching back and entering the otp ***********")
        self.driver.switch_to.default_content()

        self.driver.switch_to.window(self.driver.window_handles[0])
        self.driver.get(Url)

        self.sp = companySignUpPage(self.driver)
        # time.sleep(5)

        self.sp.setCompanyName(company_name)

        self.sp.setSearchIndustryType(setSearchIndustryType)
        self.sp.selectCompany()
        self.sp.setContactName(first_name)
        self.sp.setEmail(email2)
        self.sp.clickcountrydd()
        self.sp.clickindia()
        workbook = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        worksheet = workbook.active
        email = worksheet["E19"].value

        self.sp.clickstatedd()
        self.sp.clickTelangana()
        self.sp.clickcitydd()
        self.sp.clickHyderabad()
        self.sp.setPhone(phone_number)
        self.sp.setPassword(self.password)
        self.sp.setConfirmPassword(self.password)
        self.sp.clicktermsConditions()
        time.sleep(2)
        self.logger.info("******** Clicking on signup button ***********")
        self.logger.info("******** TC1_2  Verify that a User can Successfully Sign Up with OTP  ***********")
        self.sp.clicksignupNow()
        time.sleep(2)

        # Execute JavaScript to open a new tab
        self.driver.execute_script("window.open('about:blank', '_blank');")

        # Perform actions in the new tab (if needed)
        # For example:
        self.driver.switch_to.window(self.driver.window_handles[1])
        self.logger.info("******** Opening new url in another tab for Email OTP ***********")
        time.sleep(1)
        self.driver.get("http://mailcatch.com/en/disposable-email")
        time.sleep(1)
        yopmail = self.driver.find_element(By.XPATH, "//input[@name='box']")
        yopmail.send_keys(email2 + Keys.ENTER)
        time.sleep(1)

        reload_button = self.driver.find_element(By.XPATH, "//img[@title='Reload']")

        # Click the Reload button every second until the subject is displayed or a maximum time is reached
        max_wait_time = 60  # Set your maximum wait time in seconds
        start_time = time.time()

        while time.time() - start_time < max_wait_time:
            reload_button.click()

            try:
                # Check if the subject is displayed
                subject = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[@class='subject']"))
                )
                subject.click()
                break  # Break out of the loop if subject is displayed
            except StaleElementReferenceException:
                print("StaleElementReferenceException occurred. Retrying...")
                continue  # Retry the loop if StaleElementReferenceException occurs
            except TimeoutException:
                time.sleep(1)

        iframeElement = self.driver.find_element(By.ID, "emailframe")
        self.driver.switch_to.frame(iframeElement)

        # Code outside the loop will be executed after the loop or when a TimeoutException occurs
        otp = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//body"))
        )
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        confirmation_code = otp.text
        getOTP = re.search(r'\b\d+\b', confirmation_code).group()
        print(getOTP)

        # This code is for QA ENV
        otp = self.driver.find_element(By.XPATH, "//body")
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        confirmation_code = otp.text
        getOTP = re.search(r'\b\d+\b', confirmation_code).group()
        print(getOTP)

        self.logger.info("******** Switching back and entering the otp ***********")
        self.driver.switch_to.default_content()

        self.driver.switch_to.window(self.driver.window_handles[0])

        self.sp.setOtp(getOTP)

        time.sleep(2)
        self.logger.info(
            "******** TC3_1 Verify the Signup page OTP page,  Verify that a user can successfully verify their account with a valid OTP. ***********")
        self.sp.clickVerifyButton()
        self.sp.clickContinueToLogin()
        self.logger.info("******** Company Sign Up successful ***********")
        self.logger.info("******** Entering the sig up credentials for Login ***********")

        # Read data from specific cells

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(email2)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.lp.clickcreatePost()
        self.logger.info("******** Login successful ***********")
        act_Text = self.lp.newsFeedText()

        if act_Text == "Create News Feed":
            assert True
            self.logger.info("********* SignUp Test is Passed ***********")

        else:
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_SignUpwithValid.png")
            self.logger.error("********* SignUp Test is Failed ***********")
            self.driver.close()
            assert False

        time.sleep(3)
        self.driver.find_element(By.XPATH, "//div[@class='flexAutoRow alignCntr pdngHXS']").click()
